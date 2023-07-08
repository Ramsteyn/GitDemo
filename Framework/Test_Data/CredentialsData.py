import openpyxl
class credentialsData:

        #test_Data= [{"Firstname": "Ram", "Lastname": "Mani", "Mail": "rdsrdx49@gmail.com", "Password": "Testing123"}]

        #@staticmethod
        def cred(test_name):
                book = openpyxl.load_workbook("C:\\Users\\ramadhma\\PycharmProjects\\pythonProject1\\Excel.xlsx")
                sheet = book.active
                # print(sheet.cell(row = 2, column= 1).value)
                Dic = {}
                for i in range(1, sheet.max_row + 1):
                        if sheet.cell(row=i, column=1).value == "TC1":
                                for j in range(2, sheet.max_column + 1):
                                        # print(sheet.cell(row= i, column= j).value)
                                        Dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                return [Dic]

